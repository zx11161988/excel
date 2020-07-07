#! /user/bin/env python
from openpyxl import load_workbook


class ReadData:

    def __init__(self, tableName, sheetName):
        self.workbook = load_workbook(tableName)
        self.sheet = self.workbook[sheetName]
        self.max_row = self.sheet.max_row
        self.max_col = self.sheet.max_column
        self.list_key_value = []
        self.dict = {}
        self.dict.clear()
        print("self.max_col", self.max_col)
    #获取key列，value列，并且组装返回
    def _readData(self, index_key, index_value):
        list_key = []
        list_value = []
        print("读取数据", index_key, index_value)
        for row in range(2, int(self.max_row +1)):
            key = self.sheet.cell(row=row, column=int(index_key)).value
            value = self.sheet.cell(row=row, column=int(index_value)).value
            print("读取数据 key value", key, value)
            if(None != key):
                list_key.append(key)
                if (None == value):
                    list_value.append(0)
                else:
                    list_value.append(value)
            else:
                print("读取数据出现错误请检查 key", key)
        print("读取数据 list:", list_key)
        print("读取数据 list:", list_value)
        self.list_key_value.append(list_key)
        self.list_key_value.append(list_value)
        print("读取数据 list:", self.list_key_value)
        return self.list_key_value

    def read(self, key=None,value=None):
        print("read >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
        self._readData(key, value)
        print("read <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")
    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    def get_cell_index(self, tile):
        for col in range(1, int(self.max_col+1)):
            cell = self.sheet.cell(row=1, column=col)
            print("get_cell_index = ", cell.value)
            if(tile == cell.value):
                return cell.column
    def list2dict(self, list):
        print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
        if(None != list):
            if(len(list[0]) != len(list[1])):
                print("key的个数和value的个数不相同！！");
            else:
                for i in range(0,len(list[0])):
                    key = list[0][i]
                    value = list[1][i]
                    print("key  value ",key, value);
                    if (key in self.dict):
                        print("self.dict[key]  ", self.dict[key]);
                        self.dict[key] += value
                        print("self.dict[key]  ", self.dict[key] );
                    else:
                        self.dict[key] = value
            #for item in list[0]:
            #    print(item)
            #for item in list[1]:
            #    print(item)
        print(self.dict)
        print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>22")
if __name__ == '__main__':
    data = ReadData('AP10设备清单_2018.xlsx', '手机')
    print("max_col = ", data.max_col)
    index_key = data.get_cell_index('Owner')
    print("get index key = ", index_key)
    index_value = data.get_cell_index('测试')
    print("get index_value = ", index_value)
    list = data.read(index_key, index_value)
    data.list2dict(data.list_key_value)
